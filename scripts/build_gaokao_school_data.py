import json
import hashlib
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

SOURCE = r"E:\BaiduNetdiskDownload\6月25更河北高级版\1_2025年河北高级版(1).xlsx"
OUTPUT = Path("data/gaokao-schools.json")
JS_OUTPUT = Path("data/gaokao-schools.js")


def clean(value):
    if value is None:
        return ""
    return str(value).strip().replace("\n", "")


def number(value, default=0):
    try:
        if value in (None, "", "—", "-"):
            return default
        return int(float(value))
    except (TypeError, ValueError):
        return default


def has_flag(text, flag):
    return flag in clean(text)


def school_level_score(level):
    text = clean(level)
    score = 0
    if "985" in text:
        score += 45
    if "211" in text:
        score += 25
    if "双一流" in text:
        score += 20
    if "保研资格" in text:
        score += 6
    if "卓越" in text:
        score += 4
    return score


def school_logo_text(name):
    name = clean(name)
    if not name:
        return "校"
    for suffix in ("大学", "学院", "学校"):
        if suffix in name:
            return name[:2]
    return name[:2]


def rank_label(score, rank):
    if score >= 650 or rank <= 3000:
        return "冲刺"
    if score >= 590 or rank <= 25000:
        return "稳妥"
    return "保底"


def main():
    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    ws = wb["最终版"]
    schools = {}
    province_counts = defaultdict(set)

    for row in ws.iter_rows(min_row=4, values_only=True):
        school_name = clean(row[13])
        if not school_name:
            continue
        province = clean(row[55])
        city = clean(row[56])
        owner = clean(row[58])
        school_nature = clean(row[59])
        city_level = clean(row[60])
        doctor_points = number(row[66])
        master_points = number(row[67])
        level = clean(row[68])
        soft_rank = number(row[69])
        address = clean(row[70])
        admission_rules_url = clean(row[71])
        baike_url = clean(row[72])
        school_key = f"{province}:{school_name}"
        score_2024 = number(row[19])
        rank_2024 = number(row[20])
        major_name = clean(row[15])
        first_class_major = clean(row[3])
        major_plan = number(row[18])
        subject = clean(row[11])
        requirement = clean(row[16])
        tuition = clean(row[8]) or clean(row[7])

        province_counts[province].add(school_key)
        item = schools.setdefault(school_key, {
            "id": f"s-{hashlib.md5(school_key.encode('utf-8')).hexdigest()[:12]}",
            "code": clean(row[12]),
            "name": school_name,
            "logoText": school_logo_text(school_name),
            "province": province,
            "city": city,
            "nature": clean(row[57]),
            "owner": owner,
            "schoolNature": school_nature,
            "cityLevel": city_level,
            "doctorPoints": doctor_points,
            "masterPoints": master_points,
            "level": level,
            "softRank": soft_rank,
            "address": address,
            "admissionRulesUrl": admission_rules_url,
            "baikeUrl": baike_url,
            "is985": has_flag(level, "985"),
            "is211": has_flag(level, "211"),
            "isDoubleFirst": has_flag(level, "双一流"),
            "hasPostgraduate": has_flag(level, "保研资格"),
            "minScore": score_2024 or 999,
            "minRank": rank_2024 or 99999999,
            "maxScore": score_2024,
            "majorCount": 0,
            "planTotal": 0,
            "subjects": set(),
            "requirements": set(),
            "nationalFirstClassMajors": set(),
            "majors": [],
            "scoreLevel": school_level_score(level),
        })

        if owner and not item["owner"]:
            item["owner"] = owner
        if school_nature and not item["schoolNature"]:
            item["schoolNature"] = school_nature
        if city_level and not item["cityLevel"]:
            item["cityLevel"] = city_level
        if doctor_points:
            item["doctorPoints"] = max(item["doctorPoints"], doctor_points)
        if master_points:
            item["masterPoints"] = max(item["masterPoints"], master_points)
        if soft_rank:
            item["softRank"] = item["softRank"] or soft_rank
        if address and not item["address"]:
            item["address"] = address
        if admission_rules_url and not item["admissionRulesUrl"]:
            item["admissionRulesUrl"] = admission_rules_url
        if baike_url and not item["baikeUrl"]:
            item["baikeUrl"] = baike_url

        if score_2024:
            item["minScore"] = min(item["minScore"], score_2024)
            item["maxScore"] = max(item["maxScore"], score_2024)
        if rank_2024:
            item["minRank"] = min(item["minRank"], rank_2024)
        item["majorCount"] += 1
        item["planTotal"] += major_plan
        if subject:
            item["subjects"].add(subject)
        if requirement:
            item["requirements"].add(requirement)
        if major_name and "国家级" in first_class_major:
            item["nationalFirstClassMajors"].add(major_name.split("[")[0].strip())
        if major_name and len(item["majors"]) < 8:
            item["majors"].append({
                "code": clean(row[14]),
                "name": major_name,
                "score": score_2024,
                "rank": rank_2024,
                "plan": major_plan,
                "tuition": tuition,
                "requirement": requirement,
                "tag": rank_label(score_2024, rank_2024),
            })

    normalized_schools = []
    for item in schools.values():
        min_score = 0 if item["minScore"] == 999 else item["minScore"]
        min_rank = 0 if item["minRank"] == 99999999 else item["minRank"]
        item["minScore"] = min_score
        item["minRank"] = min_rank
        item["subjects"] = sorted(item["subjects"])
        item["requirements"] = sorted(item["requirements"])
        item["nationalFirstClassMajors"] = sorted(item["nationalFirstClassMajors"])[:12]
        item["recommendScore"] = item["scoreLevel"] + min(40, max(0, (item["maxScore"] - 500) / 5)) + min(20, item["majorCount"] / 8)
        item["guaranteeRate"] = max(8, min(96, round(100 - (item["maxScore"] - 520) * 0.22)))
        item["researchLevel"] = "A+" if item["is985"] else ("A" if item["is211"] or item["isDoubleFirst"] else "B+")
        item["employment"] = f"{max(6, min(28, round((item['majorCount'] + item['planTotal']) / 8, 1)))}K"
        normalized_schools.append(item)

    normalized_schools.sort(key=lambda x: (-x["recommendScore"], x["minRank"] or 99999999, -x["maxScore"]))
    for index, item in enumerate(normalized_schools, 1):
        item["rank"] = index
        item["match"] = max(68, min(98, round(99 - index * 0.16)))
        item.pop("scoreLevel", None)

    provinces = [
        {"name": province, "count": len(keys)}
        for province, keys in province_counts.items()
        if province
    ]
    provinces.sort(key=lambda x: (-x["count"], x["name"]))

    payload = {
        "source": "1_2025年河北高级版(1).xlsx",
        "examProvince": "河北省",
        "updatedAt": "2026-06-07",
        "summary": {
            "schoolCount": len(normalized_schools),
            "provinceCount": len(provinces),
            "doubleFirstCount": sum(1 for item in normalized_schools if item["isDoubleFirst"]),
            "count985": sum(1 for item in normalized_schools if item["is985"]),
            "count211": sum(1 for item in normalized_schools if item["is211"]),
        },
        "provinces": provinces,
        "schools": normalized_schools[:420],
    }
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    text = json.dumps(payload, ensure_ascii=False, indent=2)
    OUTPUT.write_text(text, encoding="utf-8")
    JS_OUTPUT.write_text(f"window.gaokaoSchoolData = {text};\n", encoding="utf-8")
    print(f"wrote {OUTPUT} and {JS_OUTPUT} schools={len(payload['schools'])} provinces={len(provinces)}")
    print(payload["summary"])


if __name__ == "__main__":
    main()
