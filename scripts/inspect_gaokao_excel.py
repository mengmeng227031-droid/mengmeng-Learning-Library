from openpyxl import load_workbook

SOURCE = r"E:\BaiduNetdiskDownload\6月25更河北高级版\1_2025年河北高级版(1).xlsx"


def main():
    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    print(wb.sheetnames)
    if "最终版" in wb.sheetnames:
        ws = wb["最终版"]
        header_rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
        print("\n## 最终版 headers")
        for index, values in enumerate(zip(*header_rows), 1):
            print(index, values)
    for ws in wb.worksheets[:8]:
        print(f"\n## {ws.title} rows={ws.max_row} cols={ws.max_column}")
        for row in ws.iter_rows(min_row=1, max_row=min(8, ws.max_row), values_only=True):
            print(list(row[:30]))


if __name__ == "__main__":
    main()
